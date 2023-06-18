import openpyxl

def getRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_row)

def getcolumnCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)

def readData(file,sheetname,rownum,coloumnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(rownum,coloumnno).value

def writeData(file,sheetname,rownum,coloumnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownum,coloumnno).value=data
    workbook.save(file)

